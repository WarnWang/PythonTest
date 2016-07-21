#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: format_all_info
# Author: Mark Wang
# Date: 1/7/2016

import openpyxl
from tabulate import tabulate

sheet_name_dict = {'Artificial_Neural_Network'.lower(): "ANN",
                   'Artificial_Random'.lower(): 'ANN + Random Forest',
                   'Linear_Logistic'.lower(): 'Linear Regression + Logistic Regression',
                   'Linear_Regression'.lower(): 'Linear Regression',
                   'Random_Forest'.lower(): 'Random Forest',
                   'Random_SVM'.lower(): 'Random Forest + SVM'}


def format_CDC_MACD(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.get_sheet_names():
        if sheet_name.lower() not in sheet_name_dict:
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        row_len = len(ws.rows)
        for i in range(2, row_len + 1):
            ws['C{}'.format(i)] = float(ws['C{}'.format(i)].value) / 100
            ws['F{}'.format(i)] = float(ws['F{}'.format(i)].value) / 100

    wb.save(file_path)


def transform_to_latex(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    headers = ['Symbol', 'MSE', 'MAPE', 'MAD', 'RMSE', 'CDC', 'HMSE', 'ME']
    table_info = {}
    for sheet_name in wb.get_sheet_names():
        if sheet_name.lower() not in sheet_name_dict:
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        row_len = ws.max_row
        tables = []
        for i in range(2, row_len + 1):
            tables.append([ws['A{}'.format(i)].value,
                           ws['B{}'.format(i)].value,
                           '%.2f%%' % (float(ws['C{}'.format(i)].value) * 100),
                           ws['D{}'.format(i)].value,
                           ws['E{}'.format(i)].value,
                           '%.2f%%' % (float(ws['F{}'.format(i)].value) * 100),
                           ws['G{}'.format(i)].value,
                           ws['H{}'.format(i)].value,
                           ])
        table_info[sheet_name] = tabulate(tables, headers=headers, tablefmt='latex')

    with open('table.tex', 'w') as f:
        for algorithm in table_info:
            f.write('''\\begin{table}[h]
\\centering
\\resizebox{\\textwidth}{!}{%
''')
            f.write(table_info[algorithm])
            f.write(r'''
}
\caption{%s}
\end{table}''' % sheet_name_dict[algorithm.lower()])
            f.write('\n\n\n')


if __name__ == '__main__':
    path = '/Users/warn/Documents/Projects/stock_price/output_data/output3_2012_2015/all_info.xlsx'
    format_CDC_MACD(path)
    transform_to_latex(file_path=path)
