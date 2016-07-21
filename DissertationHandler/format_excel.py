#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: format_excel
# Author: Mark Wang
# Date: 7/7/2016

import csv
import os

import openpyxl

folder_path = '/Users/warn/Documents/Projects/stock_price/adj_close/output_1_2_2013_2016/'

already_formatted = {'0033.HK', '0069.HK', '0076.HK', '0078.HK', '0083.HK', '0101.HK', '0116.HK'}


def format_excel(path):
    file_path = os.path.join(folder_path, 'all_info.xlsx')

    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.get_sheet_names():
        if sheet_name in {'CDC', 'MAPE'}:
            continue

        ws = wb.get_sheet_by_name(sheet_name)

        for row in ws.rows[1:]:
            if str(row[0].value) in already_formatted:
                continue
            row[2].value /= 100
            row[5].value /= 100

    wb.save(file_path)


def format_csv(path):
    for file_path, dirs, file_list in os.walk(path):
        if 'stock_info.csv' not in file_list:
            continue

        csv_path = os.path.join(file_path, 'stock_info.csv')
        stock_list = []
        headers = None

        with open(csv_path) as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for line in reader:
                if line['stock'] not in already_formatted:
                    line['CDC'] = float(line['CDC']) / 100
                    line['MAPE'] = float(line['MAPE']) / 100
                stock_list.append(line)

        with open(csv_path, 'w') as f:
            writer = csv.DictWriter(f, headers)
            writer.writeheader()
            writer.writerows(stock_list)


def add_mape_cdc_mad_mse_data(path):
    name_dict = {'linear': 'Linear Regression',
                 'logistic': "Logistic Regression",
                 'random': "Random Forest",
                 'artificial': "ANN",
                 'svm': "SVM",
                 }
    wb = openpyxl.load_workbook(path)
    required_symbol = ['0005.HK', '0014.HK', '0019.HK', '0023.HK', '0031.HK', '0043.HK', '0291.HK', '0700.HK',
                       '2383.HK', '6823.HK']
    method_info_dict = {}
    sheet_name_list = wb.get_sheet_names()
    for sheet_name in sheet_name_list:
        method_list = sheet_name.split('_')
        if method_list[0] not in name_dict:
            continue

        new_name = name_dict[method_list[0]]
        if method_list[-1] in name_dict:
            new_name = '{} + {}'.format(new_name, name_dict[method_list[-1]])
        method_info_dict[new_name] = {'MAPE': [], 'MAD': [], 'MSE': [], 'CDC': [], 'AVG': []}
        ws = wb.get_sheet_by_name(sheet_name)
        for row in ws.rows:
            if row[0].value in required_symbol:
                method_info_dict[new_name]['MAPE'].append(row[2].value)
                method_info_dict[new_name]['MAD'].append(row[3].value)
                method_info_dict[new_name]['MSE'].append(row[1].value)
                method_info_dict[new_name]['CDC'].append(row[5].value)
                method_info_dict[new_name]['AVG'].append(row[-1].value)

    method_list = method_info_dict.keys()

    def add_sheet_to_workbook(sheet_name):
        if sheet_name in sheet_name_list:
            wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
        ws = wb.create_sheet(sheet_name)

        header = ['Symbol']
        header.extend(method_list)
        header.append('Average Price')
        ws.append(header)

        for i in range(len(required_symbol)):
            row = [required_symbol[i]]
            for method in method_list:
                row.append(method_info_dict[method][sheet_name][i])
            row.append(method_info_dict[method_list[0]]['AVG'][i])
            ws.append(row)

    add_sheet_to_workbook('MAPE')
    add_sheet_to_workbook('MAD')
    add_sheet_to_workbook('MSE')
    add_sheet_to_workbook('CDC')

    wb.save(path)


if __name__ == '__main__':
    add_mape_cdc_mad_mse_data('/Users/warn/Documents/Projects/StockPrice/AdjClose/output_1_2_2013_2016/all_info.xlsx')
