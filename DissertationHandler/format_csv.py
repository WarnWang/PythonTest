#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: format_csv
# Author: Mark Wang
# Date: 27/6/2016


import os

import openpyxl

from util import *

dir_path = '/Users/warn/Documents/Projects/StockPrice'

for path, dirs, files in os.walk(dir_path):
    if 'stock_info.csv' in files:
        print path
        parent_path = os.path.join('/', '/'.join(path.split('/')[:-1]))
        # print parent_path
        all_info_file = os.path.join(parent_path, 'all_info.xlsx')
        if os.path.isfile(all_info_file):
            wb = openpyxl.load_workbook(all_info_file)
        else:
            wb = openpyxl.Workbook()
        sheet_name = path.split('/')[-1]
        if sheet_name in wb:
            ws = wb.remove_sheet(wb.get_sheet_by_name(name=sheet_name))
        ws = wb.create_sheet(title=sheet_name)
        stock_csv_info = os.path.join(path, 'stock_info.csv')
        csv_file = open(stock_csv_info)
        csv_reader = csv.DictReader(csv_file)

        ws['A1'] = 'Symbol'
        ws['B1'] = 'MSE'
        ws['C1'] = 'MAPE'
        ws['D1'] = 'MAD'
        ws['E1'] = 'RMSE'
        ws['F1'] = 'CDC'
        ws['G1'] = 'HMSE'
        ws['H1'] = 'ME'
        ws['I1'] = 'AVG.'
        rowIndex = 2

        for line in csv_reader:
            ws['A{}'.format(rowIndex)].value = line['stock']
            ws['B{}'.format(rowIndex)].value = float(line['MSE'])
            cdc = float(line['CDC'])
            if cdc > 1:
                ws['F{}'.format(rowIndex)].value = cdc / 100
                ws['C{}'.format(rowIndex)].value = float(line['MAPE']) / 100
            else:
                ws['F{}'.format(rowIndex)].value = cdc
                ws['C{}'.format(rowIndex)].value = float(line['MAPE'])

            ws['D{}'.format(rowIndex)].value = float(line['MAD'])
            ws['E{}'.format(rowIndex)].value = float(line['RMSE'])

            predict_result_path = os.path.join(path, line['stock'][:4], "predict_result.csv")

            if 'HMSE' in line and line['HMSE']:
                ws['G{}'.format(rowIndex)].value = float(line['HMSE'])
                ws['H{}'.format(rowIndex)].value = float(line['ME'])

            else:
                ws['G{}'.format(rowIndex)].value = get_hmse_from_csv(predict_result_path)
                ws['H{}'.format(rowIndex)].value = get_me_from_csv(predict_result_path)

            ws['I{}'.format(rowIndex)] = get_avg_from_csv(predict_result_path)
            ws['F{}'.format(rowIndex)].number_format = '0.00%'
            ws['C{}'.format(rowIndex)].number_format = '0.00%'

            ws['I{}'.format(rowIndex)].number_format = '_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'

            for character in ['B', 'D', 'E', 'H', 'G']:
                cell_index = '{}{}'.format(character, rowIndex)
                if ws[cell_index].value > 0.05:
                    ws[cell_index].number_format = '0.00'
                else:
                    ws[cell_index].number_format = '0.00E+00'

            rowIndex += 1
        csv_file.close()

        if 'Sheet' in wb:
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

        wb.save(all_info_file)
