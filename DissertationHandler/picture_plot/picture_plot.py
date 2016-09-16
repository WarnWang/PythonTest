#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: picture_plot
# Author: Mark Wang
# Date: 28/6/2016

import os
import random

import openpyxl

from DissertationHandler.util import *
from ForFun.test_char import test_char

sheet_name_dict = {'Artificial_Neural_Network'.lower(): "ANN",
                   'Artificial_Random'.lower(): 'ANN + Random Forest',
                   'Linear_Logistic'.lower(): 'Linear Regression + Logistic Regression',
                   'Linear_Regression'.lower(): 'Linear Regression',
                   'Linear_Random'.lower(): 'Linear Regression + Random Forest',
                   'Random_Logistic'.lower(): 'Random Forest + Logistic Regression',
                   'Random_Forest'.lower(): 'Random Forest',
                   'Artificial_SVM'.lower(): 'ANN + SVM',
                   'Random_SVM'.lower(): 'Random Forest + SVM'}

short_name_dict = {'Artificial_Neural_Network'.lower(): "ANN",
                   'Artificial_Random'.lower(): 'ANN_RF',
                   'Linear_Logistic'.lower(): 'LR_LR',
                   'Linear_Regression'.lower(): 'LR',
                   'Linear_Random'.lower(): 'LR_RT',
                   'Random_Logistic'.lower(): 'RT_LR',
                   'Random_Forest'.lower(): 'RF',
                   'Artificial_SVM'.lower(): 'ANN_SVM',
                   'Random_SVM'.lower(): 'RF_SVM'}

picture_list = ['MSE', 'MAPE', 'MAD', 'RMSE', 'CDC', 'HMSE', 'ME']
# picture_list = ['MAPE', 'HMSE']

color_list = ['r', 'two_int', 'g', 'dirs', 'orange', 'brown']
algorithm_list = ['ANN', 'Linear Regression', 'Random Forest', 'ANN + Random Forest',
                  'Linear Regression + Logistic Regression', 'Random Forest + SVM']


def find_ith_number(data_list, size):
    if len(data_list) < size:
        return min(data_list)
    tmp = random.choice(data_list)
    larger = []
    lower = []
    equal = 0
    for data in data_list:
        if data > tmp:
            larger.append(data)
        elif data < tmp:
            lower.append(data)
        else:
            equal += 1

    if len(larger) > size:
        return find_ith_number(larger, size)
    elif len(larger) + equal < size:
        return find_ith_number(lower, size - len(larger) - equal)
    else:
        return tmp


class PicturePlot(object):
    def __init__(self, xlsx_path, save_path):
        self.wb_path = xlsx_path
        self.save_path = save_path
        self.tick_list = []
        self.tick_size = 10

    def read_data_from_file(self):
        wb = openpyxl.load_workbook(self.wb_path, read_only=True)
        algorithm_info = {}

        for sheet_name in wb.get_sheet_names():
            if self.tick_list:
                add_tick = False
            else:
                add_tick = True
            if sheet_name.lower() not in sheet_name_dict:
                continue

            algorithm = sheet_name_dict[sheet_name.lower()]
            algorithm_info[algorithm] = {}
            ws = wb.get_sheet_by_name(name=sheet_name)
            for row in ws.rows:
                if row[0].value.endswith('.HK'):
                    if add_tick:
                        self.tick_list.append(row[0].value)
                    algorithm_info[algorithm][row[0].value] = {}
                    for i in range(1, len(row)):
                        algorithm_info[algorithm][row[0].value][ws["{}1".format(test_char(i))].value.strip()] = row[
                            i].value

            if add_tick:
                cdc_list = []
                for symbol in algorithm_info[algorithm]:
                    cdc_list.append(algorithm_info[algorithm][symbol]['CDC'])
                value = find_ith_number(cdc_list, self.tick_size)
                index_list = []
                new_tick_list = []
                for index in range(0, len(cdc_list)):
                    if cdc_list[index] > value:
                        index_list.append(index)
                        new_tick_list.append(self.tick_list[index])
                self.tick_list = new_tick_list

        return algorithm_info

    def get_monthly_data(self, start_date='2014-01-06'):
        import matplotlib.pyplot as plt
        test_path = self.wb_path.split('/')[:-1]
        test_path = os.path.join('/', '/'.join(test_path))
        mape_method_dict = {}
        cdc_method_dict = {}
        number_dict = {}
        ticks = range(1, 13)
        jump_set = {'0845.HK', '0872.HK', '1181.HK', '1230.HK', '1314.HK', '3777.HK', '8050.HK'}
        fig = plt.figure()
        for path, dirs, files in os.walk(test_path):
            if "predict_result.csv" in files:
                method = path.split('/')[-2]
                symbol = path.split('/')[-1]
                if '{}.HK'.format(symbol) in jump_set:
                    continue
                if method not in mape_method_dict:
                    mape_method_dict[method] = np.zeros(12)

                if method not in cdc_method_dict:
                    cdc_method_dict[method] = np.zeros(12)
                if method not in number_dict:
                    number_dict[method] = 0
                save_path = os.path.join(self.save_path, "{}_{}.png".format(short_name_dict[method.lower()], symbol))
                mapes = get_monthly_mape(os.path.join(path, "predict_result.csv"))
                cdcs = get_monthly_cdc(os.path.join(path, "predict_result.csv"))
                mape_method_dict[method] += mapes
                cdc_method_dict[method] += cdcs
                number_dict[method] += 1
                mapes = map(lambda x: x * 100, mapes)
                cdcs = map(lambda x: x * 100, cdcs)

                # plt.plot(ticks, mapes)
                # plt.ylabel("MAPE (%)")
                # plt.grid(True)
                # plt.title('Stock {}.HK monthly MAPE method {}'.format(symbol, sheet_name_dict[method.lower()]))
                # save_path = os.path.join(self.save_path,
                #                          "{}_{}_MAPE.png".format(short_name_dict[method.lower()], symbol))
                # plt.savefig(save_path)
                # plt.clf()
                #
                # plt.plot(ticks, cdcs)
                # plt.ylabel("CDC (%)")
                # plt.grid(True)
                # plt.title('Stock {}.HK monthly CDC method {}'.format(symbol, sheet_name_dict[method.lower()]))
                # save_path = os.path.join(self.save_path,
                #                          "{}_{}_CDC.png".format(short_name_dict[method.lower()], symbol))
                # plt.savefig(save_path)
                # plt.clf()

        wb_path = os.path.join(self.save_path, 'monthly_data.xlsx')
        if os.path.isfile(wb_path):
            wb = openpyxl.load_workbook(wb_path)
        else:
            wb = openpyxl.Workbook()

        if 'MAPE' in wb.get_sheet_names():
            wb.remove_sheet(wb.get_sheet_by_name('MAPE'))

        if 'CDC' in wb.get_sheet_names():
            wb.remove_sheet(wb.get_sheet_by_name('CDC'))
        mape_ws = wb.create_sheet('MAPE')
        cdc_ws = wb.create_sheet('CDC')
        headers = ['Method']
        headers.extend(range(1, 13))

        method_list = mape_method_dict.keys()
        mape_ws.append(headers)
        cdc_ws.append(headers)
        for method in method_list:
            cdc_row = [get_new_name(method)]
            mape_row = [get_new_name(method)]
            cdc_method_dict[method] /= number_dict[method]
            mape_method_dict[method] /= number_dict[method]
            cdc_row.extend(cdc_method_dict[method])
            mape_row.extend(mape_method_dict[method])
            mape_ws.append(mape_row)
            cdc_ws.append(cdc_row)

        wb.save(wb_path)

        # for method in mape_method_dict:
        # save_path = os.path.join(self.save_path, "{}_AVG_MAPE.png".format(short_name_dict[method.lower()]))
        # plt.plot(ticks, mape_method_dict[method] / number_dict[method] * 100)
        # plt.ylabel("MAPE (%)")
        # plt.grid(True)
        # plt.title('{} {} stocks monthly MAPE'.format(sheet_name_dict[method.lower()], number_dict[method]))
        # plt.savefig(save_path)
        # plt.clf()
        #
        # save_path = os.path.join(self.save_path, "{}_AVG_CDC.png".format(short_name_dict[method.lower()]))
        # plt.plot(ticks, cdc_method_dict[method] / number_dict[method] * 100)
        # plt.ylabel("CDC (%)")
        # plt.grid(True)
        # plt.title('{} {} stocks monthly CDC'.format(sheet_name_dict[method.lower()], number_dict[method]))
        # plt.savefig(save_path)
        # plt.clf()

    def get_picture(self):
        import matplotlib.pyplot as plt
        info_dict = self.read_data_from_file()
        for i, picture in enumerate(picture_list):
            fig = plt.figure(i)
            list_dict = {}
            new_tick_list = self.tick_list[:]
            for algorithm in sheet_name_dict.values():
                if algorithm in info_dict:
                    list_dict[algorithm] = []
            for symbol in new_tick_list:
                for algorithm in list_dict:
                    # if picture in ['HMSE']:
                    #     list_dict[algorithm].append(np.log10(info_dict[algorithm][symbol][picture]))
                    # elif picture == 'MAPE':
                    #     list_dict[algorithm].append(np.log10(info_dict[algorithm][symbol][picture] / 100))
                    # else:
                    list_dict[algorithm].append(info_dict[algorithm][symbol][picture])

            opacity = 0.4
            bar_width = 0.12
            index = np.arange(len(new_tick_list))

            for i, algorithm in enumerate(list_dict.keys()):
                plt.bar(index + bar_width * i, list_dict[algorithm], bar_width,
                        alpha=opacity,
                        color=color_list[i],
                        label=algorithm)
            plt.xlabel('Stock Symbol')
            plt.ylabel(picture)
            plt.xticks(index + 3 * bar_width, new_tick_list)
            plt.legend(loc=0)
            file_path = os.path.join(self.save_path, '{}.png'.format(picture))
            size = fig.get_size_inches()
            fig.set_size_inches(size * 1.5, forward=False)
            fig.savefig(file_path)

        for algorithm in info_dict:
            cdc = 0.0
            for symbol in self.tick_list:
                cdc += info_dict[algorithm][symbol]['CDC']
            print algorithm, cdc / len(self.tick_list)
        plt.close()


if __name__ == '__main__':
    test = PicturePlot(
        xlsx_path='/Users/warn/Documents/Projects/StockPrice/AdjClose/output_1_2_2013_2016/all_info.xlsx',
        save_path='/Users/warn/Documents/Projects/DissertationReport/graduate-thesis/Figures/AdjClose/20132016/Monthly')

    test.get_monthly_data(start_date='2015-01-06')
    # test.get_picture()
    # import pprint

    # info = test.read_data_from_file()
    # pprint.pprint(info['Random Forest + SVM'], width=150)
    # pprint.pprint(info['Linear Regression + Logistic Regression'], width=150)
    #
    # for algorithm in info:
    #     mape = 0.0
    #     for stock_symbol in info[algorithm]:
    #         mape += info[algorithm][stock_symbol]['MAPE']
    #     print algorithm, mape / 10
