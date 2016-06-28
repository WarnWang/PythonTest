#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: format_csv
# Author: Mark Wang
# Date: 27/6/2016


import os

import openpyxl

from ForFun.test_char import test_char

dir_path = '/Users/warn/PycharmProjects/output_data'

for x, y, z in os.walk(dir_path):
    if 'output3_2012_2015' not in x:
        continue
    if 'stock_info.csv' in z:
        # print x, y, z
        parent_path = os.path.join('/', '/'.join(x.split('/')[:-1]))
        # print parent_path
        all_info_file = os.path.join(parent_path, 'all_info.xlsx')
        if os.path.isfile(all_info_file):
            wb = openpyxl.load_workbook(all_info_file)
        else:
            wb = openpyxl.Workbook()
        sheet_name = x.split('/')[-1]
        if sheet_name in wb:
            ws = wb.remove_sheet(wb.get_sheet_by_name(name=sheet_name))
        ws = wb.create_sheet(title=sheet_name)
        csv_info = os.path.join(x, 'stock_info.csv')
        csv_reader = open(csv_info)
        rowIndex = 1

        for line in csv_reader:
            colIndex = 0
            for info in line.split(','):
                cellIndex = "{}{}".format(test_char(colIndex), rowIndex)
                try:
                    info = float(info)
                except Exception, err:
                    print err
                ws[cellIndex] = info
                colIndex += 1

            rowIndex += 1
        csv_reader.close()
        if 'Sheet' in wb:
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        wb.save(all_info_file)
