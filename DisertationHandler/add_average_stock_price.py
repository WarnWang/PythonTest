#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Project: PythonTest
# File name: add_average_stock_price
# Author: Mark Wang
# Date: 28/6/2016


import os

import openpyxl

from util import *

dir_path = '/Users/warn/PycharmProjects/output_data'

for x, y, z in os.walk(dir_path):
    if 'all_info.xlsx' in z:
        all_info_file = os.path.join(x, 'all_info.xlsx')
        wb = openpyxl.load_workbook(all_info_file)
        for sheet_name in wb.get_sheet_names():
            ws = wb.get_sheet_by_name(name=sheet_name)
            ws['I1'] = "AVG."
            for i in range(2, 12):
                stock_name = ws['A{}'.format(i)].value
                predict_result_path = os.path.join(x, sheet_name, stock_name[:4], "predict_result.csv")
                if not os.path.isfile(predict_result_path):
                    continue
                avg = get_avg_from_csv(predict_result_path)
                ws['I{}'.format(i)] = avg

        wb.save(all_info_file)
